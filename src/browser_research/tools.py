"""Tool implementations — pure async, mirror the authority-web-search-mcp
patterns (shared client singletons, cachetools TTL, JSON timing logs).

Browser engine is patchright (a drop-in patched Playwright). One Chromium
per process; one BrowserContext per MCP client/session for cookie + auth
isolation. Pages are short-lived — open, read, close — so a single Cloud
Run instance can soak many concurrent visits.
"""
from __future__ import annotations

import asyncio
import base64
import contextvars
import json
import logging
import os
import re
import time
from datetime import datetime, timezone
from typing import Any
from urllib.parse import urlparse

import httpx
from cachetools import TTLCache

# Use patchright — drop-in async Playwright with anti-detection patches.
# All names match upstream Playwright so the swap is one import line.
from patchright.async_api import async_playwright, Browser, BrowserContext, Page

from .extraction import STRUCTURED_EXTRACT_SYSTEM_STATIC, dynamic_date_block

log = logging.getLogger("browser_research")


# ============================================================================
# Module-level singletons. The browser is shared across the process; contexts
# are per-client so cookies / storage / auth state don't leak between tenants.
# ============================================================================

_pw_instance: Any | None = None
_browser: Browser | None = None
# Set only when BROWSER_ENGINE=camoufox — the AsyncCamoufox context manager that
# owns the Firefox process + its own Playwright instance (closed in shutdown()).
_camoufox_mgr: Any | None = None
# Keyed by (client_id, proxied) so a direct context and a proxy-routed context
# coexist per tenant — most sites use the direct one; only playbook-flagged /
# explicitly-requested domains pay for the proxy.
_contexts: dict[tuple[str, bool], BrowserContext] = {}
_browser_lock = asyncio.Lock()
# Headful (non-headless) retry: an Xvfb process providing a virtual display so a
# real browser WINDOW can run on a headless host, and a lock so at most one
# headful Chromium is resident at a time (it's launched-and-closed per retry, so
# we never keep two browsers in memory). See _headful_fetch.
_xvfb_proc: Any | None = None
_headful_lock = asyncio.Lock()

_anthropic_client: Any | None = None
_anthropic_lock = asyncio.Lock()


_current_client: contextvars.ContextVar[str] = contextvars.ContextVar(
    "_br_current_client", default="anon",
)


def set_current_client(client_id: str | None) -> None:
    _current_client.set(client_id or "anon")


def _anthropic_key() -> str | None:
    return os.environ.get("ANTHROPIC_API_KEY") or None


def _anthropic_model() -> str:
    # Haiku 4.5 supports vision and handles the fixed-schema structured
    # extraction job (title/dateline/summary/key_facts/numeric_values/
    # dates/tables_summary) at ~1/3 the Sonnet cost. The screenshot+text
    # input is high-volume per call, so this is where the spend sat.
    # Override via ANTHROPIC_MODEL=claude-sonnet-4-6 on Cloud Run if a
    # specific page needs the bigger model for stubborn chart parsing.
    return os.environ.get("ANTHROPIC_MODEL", "claude-haiku-4-5")


def _tavily_key() -> str | None:
    return os.environ.get("TAVILY_API_KEY") or None


def _firecrawl_key() -> str | None:
    return os.environ.get("FIRECRAWL_API_KEY") or None


def _proxy_opts() -> dict[str, str] | None:
    """Playwright proxy dict from env, or None when unconfigured. The egress IP
    is the dominant block signal for enterprise CDNs (Akamai et al.) — a
    residential/ISP proxy is what actually clears them. Set BROWSER_PROXY_SERVER
    (e.g. http://gw.proxy.net:7000) plus optional BROWSER_PROXY_USERNAME /
    BROWSER_PROXY_PASSWORD; tools only route through it when use_proxy is on
    (explicitly or via a playbook `proxy` hint)."""
    server = os.environ.get("BROWSER_PROXY_SERVER")
    if not server:
        return None
    opts: dict[str, str] = {"server": server}
    user = os.environ.get("BROWSER_PROXY_USERNAME")
    pw = os.environ.get("BROWSER_PROXY_PASSWORD")
    if user:
        opts["username"] = user
    if pw:
        opts["password"] = pw
    return opts


def _httpx_proxy_url() -> str | None:
    """The same proxy as a single URL for httpx (download_file). Embeds auth
    when set. Returns None when no proxy is configured."""
    server = os.environ.get("BROWSER_PROXY_SERVER")
    if not server:
        return None
    user = os.environ.get("BROWSER_PROXY_USERNAME")
    pw = os.environ.get("BROWSER_PROXY_PASSWORD")
    if user and "://" in server:
        scheme, rest = server.split("://", 1)
        from urllib.parse import quote
        cred = quote(user, safe="")
        if pw:
            cred += ":" + quote(pw, safe="")
        return f"{scheme}://{cred}@{rest}"
    return server


def _browser_engine() -> str:
    """Which browser engine to launch: "chromium" (default, patchright) or
    "camoufox" (Firefox-based anti-detect; optional, see README). Unknown values
    fall back to chromium."""
    eng = os.environ.get("BROWSER_ENGINE", "chromium").strip().lower()
    return eng if eng in ("chromium", "camoufox") else "chromium"


async def _anthropic():
    if not _anthropic_key():
        return None
    global _anthropic_client
    if _anthropic_client is None:
        async with _anthropic_lock:
            if _anthropic_client is None:
                from anthropic import AsyncAnthropic
                _anthropic_client = AsyncAnthropic(api_key=_anthropic_key())
    return _anthropic_client


async def _get_browser() -> Browser:
    """Return the shared browser, launching the configured engine on first use."""
    global _browser
    if _browser is not None and _browser.is_connected():
        return _browser
    async with _browser_lock:
        if _browser is not None and _browser.is_connected():
            return _browser
        if _browser_engine() == "camoufox":
            _browser = await _launch_camoufox()
        else:
            _browser = await _launch_chromium()
        return _browser


async def _launch_chromium() -> Browser:
    """Patchright-patched Chromium — the default engine. Set BROWSER_CHANNEL=chrome
    to drive a real Google Chrome binary (must be installed in the image) for a
    genuine Chrome TLS/version fingerprint; unset uses the bundled Chromium."""
    global _pw_instance
    if _pw_instance is None:
        _pw_instance = await async_playwright().start()
    channel = os.environ.get("BROWSER_CHANNEL") or None
    # Container-friendly flags. --no-sandbox is required when running as a
    # non-root user inside Docker; patchright keeps the stealth patches active
    # regardless.
    browser = await _pw_instance.chromium.launch(
        headless=os.environ.get("HEADLESS", "true").lower() != "false",
        **({"channel": channel} if channel else {}),
        args=[
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-blink-features=AutomationControlled",
            "--disable-features=IsolateOrigins,site-per-process",
        ],
    )
    log.info("Chromium launched (patchright stealth, channel=%s)",
              channel or "bundled")
    return browser


async def _launch_camoufox() -> Browser:
    """Camoufox — a Firefox fork with engine-level fingerprint spoofing. Optional
    and lazily imported, so it's a no-op unless BROWSER_ENGINE=camoufox AND the
    package is installed (`pip install 'camoufox[geoip]'` + `python -m camoufox
    fetch`). Renders + screenshots like Chromium, so the Sonnet-vision path is
    unaffected."""
    global _camoufox_mgr
    try:
        from camoufox.async_api import AsyncCamoufox
    except ImportError as e:  # pragma: no cover - only hit when opted in
        raise RuntimeError(
            "BROWSER_ENGINE=camoufox but the 'camoufox' package isn't installed. "
            "Run `pip install 'camoufox[geoip]'` and `python -m camoufox fetch`, "
            "add Firefox's system libs to the image, then redeploy. "
            "See README -> Browser engines."
        ) from e
    headless_env = os.environ.get("HEADLESS", "true").lower()
    # "virtual" -> Camoufox auto-manages an Xvfb display (most stealthy on a
    # headless host; needs xvfb installed). "false" -> headful. else headless.
    headless: Any = ("virtual" if headless_env == "virtual"
                     else False if headless_env == "false" else True)
    # Kept minimal to stay launch-safe across versions. Tune later: humanize=True
    # (human cursor), os=..., proxy=..., geoip=True (auto-aligns tz/locale/WebGL
    # to a proxy's exit IP — enable once a residential proxy is wired).
    mgr = AsyncCamoufox(headless=headless, locale="en-IN")
    browser = await mgr.__aenter__()
    _camoufox_mgr = mgr
    log.info("Camoufox (Firefox) launched (headless=%s)", headless)
    return browser


async def _get_context(client_id: str, proxied: bool = False) -> BrowserContext:
    # A proxy was requested but none is configured → fall back to the direct
    # context (the request still goes out, just not via a proxy) so a stale
    # playbook `proxy` hint can never hard-fail a fetch.
    proxy = _proxy_opts() if proxied else None
    if proxied and proxy is None:
        proxied = False
    key = (client_id, proxied)
    if key in _contexts:
        ctx = _contexts[key]
        try:
            # Cheap liveness probe — accessing .pages on a closed context
            # raises, which is the signal to recreate it.
            _ = len(ctx.pages)
            return ctx
        except Exception:
            _contexts.pop(key, None)
    browser = await _get_browser()
    # India-default geo / language so JS that branches on locale (PPAC, RBI
    # dashboards) renders the Indian build.
    ctx_opts: dict[str, Any] = {
        "locale": "en-IN",
        "timezone_id": "Asia/Kolkata",
        "viewport": {"width": 1440, "height": 900},
        "accept_downloads": False,
    }
    if proxy is not None:
        ctx_opts["proxy"] = proxy
        log.info("context for %s routed via proxy %s", client_id,
                 proxy.get("server"))
    # Only pin a Chrome UA for the Chromium engine. Camoufox generates its own
    # coherent Firefox fingerprint at launch — forcing a Chrome UA onto it would
    # be a glaring inconsistency that defeats the point.
    if _browser_engine() != "camoufox":
        ctx_opts["user_agent"] = (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        )
    ctx = await browser.new_context(**ctx_opts)
    _contexts[key] = ctx
    return ctx


async def shutdown() -> None:
    """Graceful cleanup. Cloud Run signals SIGTERM ~10s before kill;
    server.py's lifespan hooks call this."""
    global _browser, _pw_instance, _contexts, _camoufox_mgr, _xvfb_proc
    for ctx in list(_contexts.values()):
        try:
            await ctx.close()
        except Exception:
            pass
    _contexts.clear()
    # Camoufox owns its browser + its own Playwright instance via the context
    # manager, so exit that instead of closing _browser / _pw_instance directly.
    if _camoufox_mgr is not None:
        try:
            await _camoufox_mgr.__aexit__(None, None, None)
        except Exception:
            pass
        _camoufox_mgr = None
        _browser = None
    if _browser is not None:
        try:
            await _browser.close()
        except Exception:
            pass
        _browser = None
    if _pw_instance is not None:
        try:
            await _pw_instance.stop()
        except Exception:
            pass
        _pw_instance = None
    if _xvfb_proc is not None:
        try:
            _xvfb_proc.terminate()
        except Exception:
            pass
        _xvfb_proc = None


# ============================================================================
# Structured timing log — same JSON schema as authority-web-search.
# ============================================================================

def _emit(tool: str, started: float, *, cache_hit: bool = False,
           extra: dict | None = None) -> None:
    payload = {
        "evt": "tool",
        "tool": tool,
        "duration_ms": round((time.perf_counter() - started) * 1000),
        "cache_hit": cache_hit,
    }
    if extra:
        payload.update(extra)
    log.info(json.dumps(payload))


# ============================================================================
# In-process TTL cache for visit() results so repeat reads inside an agent
# loop don't pay the full page-load cost again.
# ============================================================================

_visit_cache: TTLCache = TTLCache(maxsize=128, ttl=180)  # 3 min


def _stable_key(*parts: Any) -> tuple:
    out: list[Any] = []
    for p in parts:
        if isinstance(p, (list, tuple)):
            out.append(tuple(sorted(str(x) for x in p)))
        elif isinstance(p, dict):
            out.append(tuple(sorted((str(k), str(v)) for k, v in p.items())))
        elif p is None:
            out.append(None)
        else:
            out.append(str(p))
    return tuple(out)


def _parse_relaxed_json(body: str) -> dict[str, Any]:
    """Pull a JSON object out of an LLM response, tolerating truncated input.

    Sonnet sometimes hits max_tokens mid-list. We grab the longest balanced
    `{...}` substring and try to parse it; if that fails, we walk backward
    closing dangling brackets until it parses. Returns {} on total failure.
    """
    s = body.strip()
    if not s:
        return {}
    # Find the first `{`
    start = s.find("{")
    if start < 0:
        return {}
    s = s[start:]
    # Try the whole thing first.
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        pass
    # Walk backward from the end, trying every truncation point and patching
    # up unclosed braces / brackets / strings until something parses.
    for end in range(len(s), 0, -1):
        candidate = s[:end]
        # Close any open string literal first.
        if candidate.count('"') % 2 == 1:
            candidate += '"'
        # Close any open arrays / objects.
        opens = candidate.count("[") - candidate.count("]")
        if opens > 0:
            candidate += "]" * opens
        opens = candidate.count("{") - candidate.count("}")
        if opens > 0:
            candidate += "}" * opens
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            continue
    return {}


def _domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().lstrip("www.")
    except Exception:
        return ""


# ============================================================================
# Network capture + request-body parsing. The single move that turns a brittle
# JS-dropdown dashboard into a one-shot is: watch the XHR/fetch the page fires,
# learn the endpoint + params, then replay that endpoint directly (same-origin,
# so cookies / CSRF / referer all match). inspect_network() does the watching;
# call_api() does the replay; act() captures opportunistically so a UI step that
# times out still leaves the agent the API it would have triggered.
# ============================================================================

# Resource types worth recording — the data-bearing calls. Documents/scripts/
# images/fonts/css are noise for API discovery.
_CAPTURE_TYPES = ("xhr", "fetch")


def _parse_request_body(body: str | None) -> dict[str, Any] | None:
    """Best-effort structure a request post-body so the agent sees PARAM NAMES.

    Returns {"kind": "json"|"form"|"raw", "data": ...} or None. Param names are
    the prize — they're what you template into call_api to pull other periods
    (e.g. PPAC's financialYear / reportBy / pageId)."""
    if not body:
        return None
    b = body.strip()
    if not b:
        return None
    if b[:1] in "{[":
        try:
            return {"kind": "json", "data": json.loads(b)}
        except Exception:  # noqa: BLE001
            pass
    if "=" in b and "\n" not in b[:200]:
        try:
            from urllib.parse import parse_qsl
            pairs = parse_qsl(b, keep_blank_values=True)
            if pairs:
                return {"kind": "form", "data": dict(pairs)}
        except Exception:  # noqa: BLE001
            pass
    return {"kind": "raw", "data": b[:2000]}


class _NetworkRecorder:
    """Attach to a Page and record data-bearing requests (XHR/fetch) with their
    request bodies and a sample of each response. Reading a response body is
    async, so each is read in a background task that we await in finalize()."""

    def __init__(self, *, max_entries: int = 80, body_cap: int = 4000,
                 capture_types: tuple[str, ...] = _CAPTURE_TYPES) -> None:
        self.entries: list[dict[str, Any]] = []
        self._tasks: list[asyncio.Task] = []
        self.max_entries = max_entries
        self.body_cap = body_cap
        self.capture_types = capture_types

    def attach(self, page: Page) -> None:
        page.on("response", self._on_response)

    def _on_response(self, response: Any) -> None:  # sync Playwright callback
        try:
            if len(self.entries) >= self.max_entries:
                return
            req = response.request
            rtype = req.resource_type
            if rtype not in self.capture_types:
                return
            try:
                post = req.post_data
            except Exception:  # noqa: BLE001
                post = None
            try:
                ct = (response.headers or {}).get("content-type", "")
            except Exception:  # noqa: BLE001
                ct = ""
            entry: dict[str, Any] = {
                "method": req.method,
                "url": req.url,
                "resource_type": rtype,
                "status": getattr(response, "status", None),
                "content_type": ct.lower(),
                "request_body": _parse_request_body(post),
            }
            self.entries.append(entry)
            # Only bother reading text-ish bodies; binary/json get capped.
            if any(k in entry["content_type"]
                   for k in ("json", "text", "javascript", "xml")):
                self._tasks.append(
                    asyncio.ensure_future(self._read_body(response, entry)))
        except Exception:  # noqa: BLE001
            pass

    async def _read_body(self, response: Any, entry: dict[str, Any]) -> None:
        try:
            txt = await response.text()
            if txt:
                entry["response_sample"] = txt[: self.body_cap]
                entry["response_truncated"] = len(txt) > self.body_cap
        except Exception:  # noqa: BLE001
            pass

    async def finalize(self) -> list[dict[str, Any]]:
        if self._tasks:
            await asyncio.gather(*self._tasks, return_exceptions=True)
        return self.entries

    def discovered_api(self, limit: int = 12) -> list[dict[str, Any]]:
        """Distinct data endpoints, freshest first, deduped by
        (method, path, sorted param keys) — the agent-facing projection."""
        seen: set[tuple] = set()
        out: list[dict[str, Any]] = []
        for e in reversed(self.entries):
            try:
                path = urlparse(e["url"]).path
            except Exception:  # noqa: BLE001
                path = e["url"]
            rb = e.get("request_body") or {}
            keys = tuple(sorted((rb.get("data") or {}).keys())) \
                if isinstance(rb.get("data"), dict) else ()
            sig = (e["method"], path, keys)
            if sig in seen:
                continue
            seen.add(sig)
            out.append(e)
            if len(out) >= limit:
                break
        return out


# ============================================================================
# Fallback fetch chain. The primary path is a real Chromium (visit/act), but
# gov / enterprise CDNs (Akamai, Cloudflare, Imperva) routinely bot-block our
# Cloud Run egress IP and serve a 200-OK "Access Denied" / JS-challenge page
# instead of content. When that happens we re-fetch the SAME url from different
# infrastructure:
#   1) Tavily Extract     — cheap, fast, different egress IP. Needs TAVILY_API_KEY.
#   2) Anthropic web_fetch — server-side fetch via the Messages API. Reuses
#      ANTHROPIC_API_KEY (already required for extraction). Server-rendered HTML
#      + PDFs only — no JS rendering.
# Both return a visit()-shaped dict (plus a `source` tag) so the rest of the
# pipeline — caching, _sonnet_extract, the MCP response — is unchanged.
# ============================================================================

# CDN bot-walls return HTTP 200 with a tiny deny/challenge body. Match the
# common shapes so we fall through instead of handing the agent a useless page.
_BLOCK_MARKERS = (
    "access denied",
    "you don't have permission to access",
    "attention required",               # Cloudflare
    "just a moment",                    # Cloudflare JS challenge / DDoS-Guard
    "checking your browser",
    "enable javascript and cookies to continue",
    "request unsuccessful. incapsula",  # Imperva
    "/cdn-cgi/",                        # Cloudflare challenge assets
    "reference #",                      # Akamai deny reference id
)


def _looks_blocked(title: str, text: str) -> str | None:
    """Return a short reason string if (title, text) look like a CDN bot-wall
    or JS challenge rather than real page content, else None."""
    t = (title or "").lower()
    if any(m in t for m in ("access denied", "attention required",
                            "just a moment", "forbidden")):
        return "challenge_title"
    body = (text or "").strip()
    if len(body) < 32:
        return "empty_body"
    # Real pages occasionally mention "access denied" in prose, so only treat a
    # marker as a block when the whole body is short (deny pages are tiny).
    if len(body) < 2000:
        low = body.lower()
        for m in _BLOCK_MARKERS:
            if m in low:
                return f"marker:{m.strip()}"
    return None


# Login/registration gates on a *download or data view*. Distinct from a bot
# wall: no fetch-fallback gets you past auth, so the right move is to pivot to
# an open source (which the domain's playbook surfaces). Phrases are specific
# enough that an ordinary "Log in" nav link won't trip them.
_AUTHWALL_MARKERS = (
    "register with us to download",
    "register to download",
    "please register to",
    "login to download",
    "log in to download",
    "sign in to download",
    "login to access",
    "log in to access",
    "please log in to view",
    "subscription required",
    "subscribe to download",
    "members only",
    "registered users only",
)


def _looks_authwalled(text: str) -> str | None:
    """Return a reason if the page gates its data/download behind login or
    registration, else None."""
    low = (text or "").lower()
    for m in _AUTHWALL_MARKERS:
        if m in low:
            return f"auth_wall:{m}"
    return None


def _field(obj: Any, key: str, default: Any = None) -> Any:
    """Read `key` from an SDK object or a plain dict — web_fetch result blocks
    surface as either depending on the installed anthropic SDK version."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


async def _tavily_fetch(url: str, *, text_cap: int) -> dict[str, Any] | None:
    """Re-fetch `url` via Tavily Extract. Returns a visit()-shaped dict, or None
    when no key is set / transport fails / the body is empty."""
    key = _tavily_key()
    if not key:
        return None
    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            r = await client.post(
                "https://api.tavily.com/extract",
                headers={"Authorization": f"Bearer {key}"},
                json={"urls": [url], "extract_depth": "advanced",
                      "format": "markdown"},
            )
            r.raise_for_status()
            data = r.json()
    except Exception as e:  # noqa: BLE001
        log.warning("tavily fallback failed for %s: %s", url, str(e)[:120])
        return None
    results = data.get("results") or []
    raw = (results[0].get("raw_content") or "").strip() if results else ""
    if not raw:
        return None
    return {
        "url": url,
        "title": "",
        "domain": _domain(url),
        "text": raw[:text_cap],
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "current_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "source": "tavily",
    }


async def _anthropic_web_fetch(url: str, *, text_cap: int) -> dict[str, Any] | None:
    """Last-resort fetch via Anthropic's server-side web_fetch tool. The url is
    placed in the user turn because web_fetch only retrieves URLs already
    present in the conversation. Server-rendered HTML only (no JS); PDFs come
    back base64 and are skipped here (use download_file for those). Returns a
    visit()-shaped dict or None."""
    client = await _anthropic()
    if client is None:
        return None
    try:
        resp = await client.messages.create(
            model=_anthropic_model(),
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": (
                    "Use the web_fetch tool to fetch this URL, then reply "
                    f"'done': {url}"
                ),
            }],
            tools=[{
                "type": "web_fetch_20250910",
                "name": "web_fetch",
                "max_uses": 1,
                "max_content_tokens": 100_000,
            }],
        )
    except Exception as e:  # noqa: BLE001
        log.warning("anthropic web_fetch failed for %s: %s", url, str(e)[:120])
        return None

    for block in resp.content:
        if _field(block, "type") != "web_fetch_tool_result":
            continue
        result = _field(block, "content")
        rtype = _field(result, "type")
        if rtype == "web_fetch_tool_error":
            log.warning("anthropic web_fetch error for %s: %s", url,
                        _field(result, "error_code", "?"))
            return None
        if rtype != "web_fetch_result":
            continue
        fetched_url = _field(result, "url", url) or url
        doc = _field(result, "content")          # the document content block
        src = _field(doc, "source")
        if _field(src, "type") != "text":        # base64 PDF — not handled here
            return None
        text = (_field(src, "data", "") or "").strip()
        if not text:
            return None
        return {
            "url": fetched_url,
            "title": (_field(doc, "title", "") or "")[:300],
            "domain": _domain(fetched_url),
            "text": text[:text_cap],
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "current_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "source": "anthropic_web_fetch",
        }
    return None


async def _firecrawl_fetch(url: str, *, text_cap: int) -> dict[str, Any] | None:
    """Re-fetch `url` via Firecrawl's /v2/scrape. Firecrawl renders JS server-
    side AND routes through its own proxy pool, so it clears the JS-SPA + CDN-IP
    blocks that defeat our Cloud Run Chromium (and that Tavily's lighter render /
    web_fetch's no-JS path miss). Returns a visit()-shaped dict or None."""
    key = _firecrawl_key()
    if not key:
        return None
    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            r = await client.post(
                "https://api.firecrawl.dev/v2/scrape",
                headers={"Authorization": f"Bearer {key}",
                         "Content-Type": "application/json"},
                # NO `proxy`/`location`: they 500 with
                # ERR_TUNNEL_CONNECTION_FAILED (the geo-proxy tier isn't on this
                # plan). Firecrawl still fetches from its OWN IPs by default, so
                # it bypasses our datacenter-IP block regardless. `timeout` caps
                # the render so a hard site fails fast instead of hanging.
                json={"url": url, "formats": ["markdown"],
                      "onlyMainContent": True, "timeout": 30000},
            )
            r.raise_for_status()
            data = r.json()
    except Exception as e:  # noqa: BLE001
        log.warning("firecrawl fallback failed for %s: %s", url, str(e)[:120])
        return None
    if not data.get("success"):
        return None
    d = data.get("data") or {}
    md = (d.get("markdown") or "").strip()
    if not md:
        return None
    meta = d.get("metadata") or {}
    title = meta.get("title")
    if isinstance(title, list):
        title = title[0] if title else ""
    final_url = meta.get("url") or meta.get("sourceURL") or url
    return {
        "url": final_url,
        "title": (title or "")[:300],
        "domain": _domain(final_url),
        "text": md[:text_cap],
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "current_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "source": "firecrawl",
    }


def _ensure_xvfb() -> bool:
    """Ensure a virtual framebuffer is available so a HEADFUL browser can open a
    real window on a headless host. In production the container runs under
    `xvfb-run` (see Dockerfile CMD), so DISPLAY is already set before the
    Playwright driver starts — the fast path below just returns True. This
    in-process Xvfb start is a fallback for environments not wrapped by
    xvfb-run (e.g. local dev). Returns False if no display can be obtained, in
    which case the headful retry quietly no-ops. Best-effort, sync, called once.

    NOTE: starting Xvfb here only helps if it happens BEFORE the Playwright
    driver process is spawned; once the driver is up it has already captured its
    env. xvfb-run is therefore the reliable path on Cloud Run."""
    if os.environ.get("DISPLAY"):
        return True
    global _xvfb_proc
    if _xvfb_proc is not None and _xvfb_proc.poll() is None:
        return True
    import shutil
    import subprocess
    if not shutil.which("Xvfb"):
        log.warning("Xvfb not installed — headful retry unavailable")
        return False
    try:
        proc = subprocess.Popen(
            ["Xvfb", ":99", "-screen", "0", "1440x900x24", "-nolisten", "tcp"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        # Wait for Xvfb to create its display socket before any browser launches
        # against :99 — otherwise Chromium races the not-yet-ready display and
        # dies with "Target page, context or browser has been closed". Only set
        # DISPLAY + publish the proc once the socket exists, so a failed start
        # never leaves DISPLAY pointing at a dead display.
        for _ in range(60):                           # up to ~6s
            if os.path.exists("/tmp/.X11-unix/X99"):
                _xvfb_proc = proc
                os.environ["DISPLAY"] = ":99"
                return True
            if proc.poll() is not None:               # Xvfb died on startup
                log.warning("Xvfb exited during startup — headful unavailable")
                return False
            time.sleep(0.1)
        log.warning("Xvfb :99 not ready in time — headful unavailable")
        try:
            proc.terminate()
        except Exception:  # noqa: BLE001
            pass
        return False
    except Exception as e:  # noqa: BLE001
        log.warning("Xvfb start failed: %s", e)
        return False


async def _headful_render(url: str, *, text_cap: int) -> dict[str, Any] | None:
    """Open `url` in a real (non-headless) Chromium window under Xvfb, launched
    fresh and closed at the end so no second browser stays resident. Serialized
    by _headful_lock → at most one headful browser in memory at a time."""
    global _pw_instance
    async with _headful_lock:
        browser = None
        try:
            if _pw_instance is None:
                _pw_instance = await async_playwright().start()
            channel = os.environ.get("BROWSER_CHANNEL") or None
            browser = await _pw_instance.chromium.launch(
                headless=False,
                **({"channel": channel} if channel else {}),
                # --disable-gpu: no GPU on Cloud Run; headful otherwise tries to
                # init one and can crash the launch.
                args=["--no-sandbox", "--disable-dev-shm-usage",
                      "--disable-blink-features=AutomationControlled",
                      "--disable-gpu"],
            )
            ctx = await browser.new_context(
                user_agent=("Mozilla/5.0 (X11; Linux x86_64) "
                            "AppleWebKit/537.36 (KHTML, like Gecko) "
                            "Chrome/131.0.0.0 Safari/537.36"),
                locale="en-IN", timezone_id="Asia/Kolkata",
                viewport={"width": 1440, "height": 900},
            )
            page = await ctx.new_page()
            await page.goto(url, wait_until="domcontentloaded", timeout=25_000)
            await page.wait_for_timeout(1500)
            title = (await page.title()) or ""
            text = await page.evaluate("() => document.body.innerText || ''")
        finally:
            if browser is not None:
                try:
                    await browser.close()
                except Exception:  # noqa: BLE001
                    pass
    if _looks_blocked(title, text) or len((text or "").strip()) < 32:
        return None
    return {
        "url": url,
        "title": title[:300],
        "domain": _domain(url),
        "text": (text or "")[:text_cap],
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "current_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "source": "headful",
    }


async def _headful_fetch(url: str, *, text_cap: int) -> dict[str, Any] | None:
    """Headful (non-headless) retry rung. A real browser WINDOW trips far fewer
    headless-detection checks than the headless build (the navigator.webdriver /
    headless-UA / missing window-chrome signals anti-bot vendors probe) — so it
    clears *fingerprint*-gated sites the headless pass can't. It does NOT change
    the egress IP, so it can't beat datacenter-IP blocks; _fallback_fetch only
    invokes it for *content* blocks (challenge_title / marker:… / empty_body),
    skipping goto: network failures (ERR_*/timeout) + auth walls. Opt out with
    HEADFUL_RETRY=false. Fully bounded + guarded: returns None on disabled / no
    display / still-blocked / error / 40s timeout, so it can never hang or break
    the fallback chain."""
    if os.environ.get("HEADFUL_RETRY", "true").strip().lower() == "false":
        return None
    if not await asyncio.to_thread(_ensure_xvfb):
        return None
    try:
        return await asyncio.wait_for(
            _headful_render(url, text_cap=text_cap), timeout=40.0)
    except Exception as e:  # noqa: BLE001
        # Keep a long slice: Playwright appends the browser's startup stderr
        # (e.g. "error while loading shared libraries: lib*.so") AFTER the
        # generic first line, which is what pinpoints a missing GUI lib.
        log.warning("headful retry failed for %s: %s", url, str(e)[:1000])
        return None


async def _fallback_fetch(url: str, *, text_cap: int,
                          reason: str) -> dict[str, Any] | None:
    """Run the fallback chain for a url the Chromium path couldn't read:
    headful retry (real browser window — only for content/fingerprint blocks) →
    Firecrawl (JS render + proxy pool — most capable) → Tavily Extract →
    Anthropic web_fetch. Each rung is opt-in via its API key; returns the first
    success (visit()-shaped, tagged with `source` + `fallback_reason`) or None."""
    t0 = time.perf_counter()
    out = None
    # Headful retry FIRST — but only for *content* blocks: the page loaded yet
    # looks like a bot-wall / is JS-empty, where a real window may clear
    # fingerprint/headless gating. These reasons come from _looks_blocked
    # (challenge_title / marker:… / empty_body). Skip it for "goto:…" network
    # failures (ERR_* or timeout — headful shares our egress IP, would fail the
    # same way after a 25s wait) and auth walls (need a login, not a window).
    if reason.startswith(("challenge_title", "marker:", "empty_body")):
        out = await _headful_fetch(url, text_cap=text_cap)
    if out is None:
        out = await _firecrawl_fetch(url, text_cap=text_cap)
    if out is None:
        out = await _tavily_fetch(url, text_cap=text_cap)
    if out is None:
        out = await _anthropic_web_fetch(url, text_cap=text_cap)
    if out is not None:
        out["fallback_reason"] = reason
    _emit("fallback_fetch", t0, extra={
        "ok": out is not None,
        "via": (out or {}).get("source"),
        "reason": reason,
        "chars": len((out or {}).get("text") or ""),
    })
    return out


# ============================================================================
# visit — open a URL, return DOM text + screenshot. The atomic primitive
# everything else builds on.
# ============================================================================

async def visit(
    url: str,
    *,
    wait_for_selector: str | None = None,
    wait_extra_ms: int = 1500,
    timeout_ms: int = 45_000,
    screenshot: bool = True,
    full_page_screenshot: bool = False,
    text_cap: int = 30_000,
    return_screenshot_b64: bool = False,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Navigate to URL with a real Chromium and return its rendered state.

    Use when pdf_fetch / http_post_form / web_fetch all fail — the page is
    a SPA whose data lives in JavaScript, or behind a login, or hidden behind
    a dropdown that's not a separate URL.

    Args:
        url: The page URL.
        wait_for_selector: Optional CSS selector to await before reading the
            DOM. Useful when the data appears only after an AJAX call returns
            (e.g. wait for ".chart svg" on a chart page).
        wait_extra_ms: Extra settle time after the wait condition fires.
        timeout_ms: Hard timeout for the whole navigation.
        screenshot: Whether to capture a PNG screenshot internally. Even
            when True, the base64 is NOT returned in the response by default
            (see return_screenshot_b64) — extract() and act() use the
            captured screenshot in-process to feed Sonnet vision.
        full_page_screenshot: If True, scroll-stitches the whole page.
        text_cap: Cap on extracted text length (innerText).
        return_screenshot_b64: If True, echo the base64 PNG back in the
            response. Defaults to False because a typical screenshot is
            ~700KB-1MB and accumulating them across an agent's tool-call
            history blows the 1M-token context window. Tools or UIs that
            actually need the bytes (e.g. a browser-canvas pane) can opt in.

    Returns:
        {url, title, domain, text, screenshot_bytes?, screenshot_b64?,
         fetched_at, current_date}
    """
    t0 = time.perf_counter()
    if not url:
        return {"error": "url is required"}

    cache_key = _stable_key(
        "visit", url, wait_for_selector, full_page_screenshot, text_cap,
        use_proxy,
    )
    if cached := _visit_cache.get(cache_key):
        _emit("visit", t0, cache_hit=True,
               extra={"chars": len(cached.get("text") or "")})
        # Strip the base64 from cache hits too unless the caller wants it.
        # The cache keeps it because extract()/act() pull from the same
        # cache via _sonnet_extract.
        if return_screenshot_b64:
            return cached
        return {k: v for k, v in cached.items() if k != "screenshot_b64"}

    client_id = _current_client.get() or "anon"
    ctx = await _get_context(client_id, proxied=use_proxy)

    def _finalize_fb(fb: dict[str, Any]) -> dict[str, Any]:
        # Cache + shape a fallback result exactly like a normal visit() return.
        if fb.get("text"):
            _visit_cache[cache_key] = dict(fb)
        _emit("visit", t0, extra={"chars": len(fb.get("text") or ""),
                                   "via": fb.get("source")})
        if return_screenshot_b64:
            return fb
        return {k: v for k, v in fb.items() if k != "screenshot_b64"}

    page = await ctx.new_page()
    try:
        try:
            await page.goto(url, wait_until="domcontentloaded",
                             timeout=timeout_ms)
        except Exception as e:  # noqa: BLE001
            # Network-level failure (timeout, DNS, connection reset). Try the
            # alternate fetch paths before giving up.
            fb = await _fallback_fetch(url, text_cap=text_cap,
                                        reason=f"goto:{str(e)[:40]}")
            if fb is not None:
                return _finalize_fb(fb)
            _emit("visit", t0, extra={"error": f"goto: {str(e)[:80]}"})
            return {"error": f"navigation failed: {e}", "url": url}

        if wait_for_selector:
            try:
                await page.wait_for_selector(wait_for_selector,
                                              timeout=min(timeout_ms, 15_000))
            except Exception:
                # Not fatal — selector might just be slow; we proceed with
                # whatever's already on the page.
                pass
        else:
            # Best-effort: wait for network to idle, but don't block the
            # whole call if a single tracker pixel hangs.
            try:
                await page.wait_for_load_state("networkidle", timeout=8_000)
            except Exception:
                pass

        if wait_extra_ms > 0:
            await page.wait_for_timeout(wait_extra_ms)

        title = (await page.title()) or ""
        url_final = page.url
        try:
            text = await page.evaluate("() => document.body.innerText || ''")
        except Exception:
            text = ""

        # Bot-wall / challenge detection. The page loaded with HTTP 200 but the
        # body is a CDN deny notice, not content — re-fetch from other infra.
        block_reason = _looks_blocked(title, text)
        if block_reason:
            fb = await _fallback_fetch(url_final, text_cap=text_cap,
                                        reason=block_reason)
            if fb is not None:
                return _finalize_fb(fb)
            # Fallbacks unavailable/failed — fall through and return the blocked
            # page, but tag it (below) so it isn't mistaken for real content.

        # Scan the rendered DOM for download-shaped links. We CANNOT parse
        # these — Chromium just returns bytes and our toolkit has no Excel /
        # CSV reader on the browser-research side. But surfacing the URLs
        # lets the agent recommend the right tool downstream (e.g. the
        # excel_fetch_structured tool on authority-web-search-mcp) or hand
        # the link back to the user. Far better than the agent spinning
        # through visit() calls on a page whose data is all in attachments.
        file_links: list[dict[str, str]] = []
        try:
            file_links = await page.evaluate(r"""
                () => {
                  const exts = ['.xlsx','.xlsm','.xls','.csv','.tsv','.pdf','.zip','.7z','.docx','.pptx'];
                  const seen = new Set();
                  const out = [];
                  for (const a of document.querySelectorAll('a[href]')) {
                    const href = (a.href || '').trim();
                    if (!href || seen.has(href)) continue;
                    const path = (new URL(href, document.baseURI)).pathname.toLowerCase();
                    const ext = exts.find(e => path.endsWith(e));
                    if (!ext) continue;
                    seen.add(href);
                    const text = (a.innerText || a.textContent || '').replace(/\s+/g,' ').trim();
                    out.push({href, text: text.slice(0, 200), format: ext.slice(1)});
                    if (out.length >= 40) break;
                  }
                  return out;
                }
            """)
        except Exception as e:  # noqa: BLE001
            log.debug("file_links scan failed: %s", e)

        out: dict[str, Any] = {
            "url": url_final,
            "title": title[:300],
            "domain": _domain(url_final),
            "text": (text or "")[:text_cap],
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "current_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        }
        if block_reason:
            # Fallbacks were unavailable or also failed; surface the block so
            # the agent treats this as a wall rather than empty content.
            out["blocked"] = block_reason
        auth_wall = _looks_authwalled(text)
        if auth_wall:
            # Login/registration gate — no fetch-fallback helps; the matched
            # playbook (attached at the server layer) points to an open source.
            out["auth_wall"] = auth_wall
        if file_links:
            out["file_links"] = file_links
            # Per-format counts so the agent can scan at a glance.
            from collections import Counter
            out["file_links_summary"] = dict(
                Counter(fl["format"] for fl in file_links)
            )

        # Capture the screenshot bytes internally regardless — extract() and
        # act() depend on them for Sonnet vision. We only put the base64 into
        # the *returned* dict if the caller asked for it; the cache keeps the
        # full payload so an extract() call right after visit() doesn't pay
        # twice. The agent-facing tool response gets `screenshot_bytes` so it
        # knows a screenshot was taken without carrying the 1MB blob.
        shot_b64: str | None = None
        if screenshot:
            try:
                png = await page.screenshot(
                    type="png",
                    full_page=full_page_screenshot,
                )
                shot_b64 = base64.b64encode(png).decode("ascii")
                out["screenshot_bytes"] = len(png)
            except Exception as e:  # noqa: BLE001
                log.warning("screenshot failed: %s", e)

        if out["text"]:
            # Cache the full payload (including base64) so chained extract()/
            # act() calls can use it. The base64 is stripped from the returned
            # dict below.
            cache_entry = dict(out)
            if shot_b64:
                cache_entry["screenshot_b64"] = shot_b64
            _visit_cache[cache_key] = cache_entry

        if return_screenshot_b64 and shot_b64:
            out["screenshot_b64"] = shot_b64

        _emit("visit", t0,
               extra={"chars": len(out["text"]),
                      "shot_kb": round(out.get("screenshot_bytes", 0) / 1024)})
        return out
    finally:
        try:
            await page.close()
        except Exception:
            pass


# ============================================================================
# extract — visit + Sonnet structured extraction. Same response shape as
# authority-web-search-mcp's pdf_fetch_structured so the agent treats it
# identically.
# ============================================================================

async def act(
    url: str,
    steps: list[dict[str, Any]],
    *,
    focus: str = "",
    timeout_ms: int = 60_000,
    full_page_screenshot: bool = True,
    include_screenshot_in_response: bool = False,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Drive a real Chromium through a sequence of steps on a page, then run
    structured extraction on the final state.

    Use this when the data lives behind an interaction — a Year/Month
    dropdown that fires AJAX inline, a tab click that reveals a table, a
    "Load more" button, a sort header, a form submit. Anything that
    `visit` can't reach because the URL doesn't change.

    Steps are dicts; each one MUST have a single key naming the action:

      {"goto":   "https://…"}                       navigate
      {"click":  "selector"}                         left-click an element
      {"fill":   {"selector": "#q", "value": "x"}}   type into an input
      {"select": {"selector": "#year", "value": "2024-2025"}}
                                                     pick a <select> option
      {"press":  {"selector": "#q", "key": "Enter"}} press a key
      {"scroll": {"to": "bottom"|"top"|<int px>}}    scroll the page
      {"wait_for_selector": "selector"}              wait for it to exist
      {"wait_for_load_state": "networkidle"|"load"}  wait for nav state
      {"wait_ms": 1500}                              hard sleep
      {"screenshot": {"name": "after-select"}}       capture an intermediate
                                                     screenshot (returned in
                                                     `step_results`)

    The first non-`goto` step is preceded by an implicit `goto(url)`.
    The final extraction runs on the LAST state of the page; screenshot
    + Sonnet vision get the chart values drawn after every step finished.

    Args:
        url: The starting page URL.
        steps: Ordered list of action dicts.
        focus: Extraction focus passed to Sonnet (same as `extract`).
        timeout_ms: Per-step navigation/wait timeout.
        full_page_screenshot: Whether the final screenshot is full-page.
        include_screenshot_in_response: Echo the final screenshot back in
            the MCP response (default False — agents don't need 700KB
            blobs in their context).

    Returns:
        Same shape as `extract` plus:
          step_results: [{step_index, action, ok, error?, duration_ms}, …]
          final_url: The page URL after all steps ran.
    """
    t0 = time.perf_counter()
    if not url:
        return {"error": "url is required"}
    if not steps or not isinstance(steps, list):
        return {"error": "steps must be a non-empty list"}

    client_id = _current_client.get() or "anon"
    ctx = await _get_context(client_id, proxied=use_proxy)
    page = await ctx.new_page()
    # Record the XHR/fetch the page fires while we drive it. Even if a UI step
    # times out (a non-native JS widget, say), the captured endpoint lets the
    # agent pivot to call_api instead of giving up.
    rec = _NetworkRecorder()
    rec.attach(page)
    fetch_sink: list[dict[str, Any]] = []
    step_results: list[dict[str, Any]] = []
    try:
        # Initial navigation.
        try:
            await page.goto(url, wait_until="domcontentloaded",
                             timeout=timeout_ms)
        except Exception as e:  # noqa: BLE001
            return {"error": f"initial navigation failed: {e}",
                    "url": url, "step_results": []}

        for idx, step in enumerate(steps):
            if not isinstance(step, dict) or len(step) != 1:
                step_results.append({
                    "step_index": idx, "action": "invalid",
                    "ok": False,
                    "error": "each step must be {action: arg} with exactly one key",
                })
                continue
            action, arg = next(iter(step.items()))
            s_t0 = time.perf_counter()
            try:
                await _run_step(page, action, arg, timeout_ms, sink=fetch_sink)
                step_results.append({
                    "step_index": idx, "action": action, "ok": True,
                    "duration_ms": round((time.perf_counter() - s_t0) * 1000),
                })
            except Exception as e:  # noqa: BLE001
                step_results.append({
                    "step_index": idx, "action": action, "ok": False,
                    "duration_ms": round((time.perf_counter() - s_t0) * 1000),
                    "error": str(e)[:200],
                })
                # Don't bail — the agent may have provided a permissive script.

        # Final state.
        final_url = page.url
        title = (await page.title()) or ""
        try:
            text = await page.evaluate("() => document.body.innerText || ''")
        except Exception:
            text = ""

        # If the final page is a CDN bot-wall, the interaction results are
        # meaningless. Re-fetch the URL from other infra so the agent still gets
        # the page's content — but flag clearly that the steps were NOT applied
        # (a plain fetch can't replay dropdown/click/filter actions).
        block_reason = _looks_blocked(title, text)
        if block_reason:
            fb = await _fallback_fetch(final_url, text_cap=20_000,
                                        reason=block_reason)
            if fb is not None and fb.get("text"):
                out = await _sonnet_extract(fb, focus=focus)
                out["step_results"] = step_results
                out["final_url"] = fb.get("url", final_url)
                out["degraded"] = (
                    f"Live page was blocked ({block_reason}); the interaction "
                    f"steps could NOT be applied. Returned a static fetch of the "
                    f"URL via {fb.get('source')} — any dropdown/click/filter "
                    f"results are not reflected. The target site must be "
                    f"reachable from the browser to capture post-interaction state."
                )
                _emit("act", t0, extra={"steps": len(steps),
                                         "via": fb.get("source"),
                                         "degraded": True})
                return out
            # Fallbacks unavailable/failed — return the blocked extraction but
            # tag it (below).

        try:
            png = await page.screenshot(type="png",
                                          full_page=full_page_screenshot)
            shot_b64 = base64.b64encode(png).decode("ascii")
            shot_bytes = len(png)
        except Exception as e:
            log.warning("act screenshot failed: %s", e)
            shot_b64 = None
            shot_bytes = 0

        # Run the same Sonnet extraction as extract() — reuse the helper
        # by hand-constructing the `visited` dict shape.
        synthetic_visited = {
            "url": final_url,
            "title": title[:300],
            "domain": _domain(final_url),
            "text": (text or "")[:20_000],
            "screenshot_b64": shot_b64,
            "screenshot_bytes": shot_bytes,
            "fetched_at": datetime.now(timezone.utc).isoformat(),
        }
        out = await _sonnet_extract(synthetic_visited, focus=focus)
        out["step_results"] = step_results
        out["final_url"] = final_url
        if block_reason:
            out["blocked"] = block_reason
        auth_wall = _looks_authwalled(text)
        if auth_wall:
            out["auth_wall"] = auth_wall
        # Network capture → adaptive recovery. Surface the data endpoints the
        # page hit, and if a UI step failed but the page still fired a
        # data-bearing request, tell the agent to replay it via call_api rather
        # than re-driving the widget.
        await rec.finalize()
        observed = rec.discovered_api(limit=8)
        if observed:
            out["observed_api"] = [{
                "method": e["method"], "url": e["url"],
                "request_params": (e.get("request_body") or {}).get("data")
                if isinstance(e.get("request_body"), dict) else None,
                "status": e.get("status"),
            } for e in observed]
        hint = _recovery_hint(step_results, observed)
        if hint:
            out["recovery_hint"] = hint
        if fetch_sink:
            out["fetch_results"] = fetch_sink
        if include_screenshot_in_response and shot_b64:
            out["screenshot_b64"] = shot_b64
        _emit("act", t0,
               extra={"steps": len(steps), "shot_kb": round(shot_bytes / 1024),
                      "observed_api": len(observed)})
        return out
    finally:
        try:
            await page.close()
        except Exception:
            pass


def _recovery_hint(step_results: list[dict[str, Any]],
                   observed: list[dict[str, Any]]) -> str | None:
    """If a UI step failed (e.g. select_option timed out on a non-native JS
    widget) but the page still fired a data-bearing XHR/fetch, point the agent
    at the call_api replay path instead of re-driving the widget."""
    failed = [s for s in step_results
              if not s.get("ok") and s.get("action") in
              ("select", "click", "fill", "press")]
    if not failed:
        return None
    posts = [e for e in observed
             if e.get("method") in ("POST", "GET")
             and (e.get("request_body") or e.get("method") == "POST")]
    target = posts[0] if posts else (observed[0] if observed else None)
    if not target:
        return None
    rb = target.get("request_body") or {}
    params = rb.get("data") if isinstance(rb, dict) else None
    return (
        f"A UI step ({failed[0].get('action')}) failed — likely a non-native JS "
        f"widget. The page nonetheless fired {target.get('method')} "
        f"{target.get('url')}"
        + (f" with params {params}" if params else "")
        + ". Replay it directly with call_api (templating the params for the "
        "period you want); this also reaches values the widget never exposes."
    )


async def _run_step(page: Page, action: str, arg: Any, timeout_ms: int,
                    sink: list[dict[str, Any]] | None = None) -> None:
    """Dispatch a single step from act()'s steps[] to Playwright. Each branch
    is intentionally narrow — anything else is rejected so the agent learns
    the supported vocabulary. `sink` collects the return value of data-yielding
    steps (currently fetch_json) so act() can surface them."""
    bounded = lambda v, default: min(int(v or default), timeout_ms)  # noqa: E731

    if action == "fetch_json":
        # In-page fetch from the page's own origin — cookies / CSRF / referer
        # all match, so a same-origin AJAX endpoint replays cleanly. THE move
        # for JS-dropdown dashboards: skip the widget, hit the endpoint it fires.
        if not isinstance(arg, dict) or not arg.get("url"):
            raise ValueError("fetch_json requires {url, method?, body?, headers?}")
        res = await _page_fetch(
            page, arg["url"], method=str(arg.get("method", "GET")),
            body=arg.get("body"), headers=arg.get("headers"),
            content_type=arg.get("content_type"),
        )
        if sink is not None:
            sink.append({"request": {k: arg.get(k) for k in
                                     ("url", "method", "body")}, "result": res})
        return
    if action == "goto":
        await page.goto(str(arg), wait_until="domcontentloaded",
                          timeout=timeout_ms)
    elif action == "click":
        await page.click(str(arg), timeout=bounded(None, 15_000))
    elif action == "fill":
        if not isinstance(arg, dict):
            raise ValueError("fill requires {selector, value}")
        await page.fill(arg["selector"], str(arg.get("value", "")),
                          timeout=bounded(None, 15_000))
    elif action == "select":
        if not isinstance(arg, dict):
            raise ValueError("select requires {selector, value}")
        await page.select_option(arg["selector"], str(arg["value"]),
                                    timeout=bounded(None, 15_000))
    elif action == "press":
        if not isinstance(arg, dict):
            raise ValueError("press requires {selector, key}")
        await page.press(arg["selector"], str(arg["key"]),
                          timeout=bounded(None, 15_000))
    elif action == "scroll":
        to = (arg or {}).get("to") if isinstance(arg, dict) else arg
        if to in ("bottom", "end"):
            await page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
        elif to in ("top", "start"):
            await page.evaluate("() => window.scrollTo(0, 0)")
        else:
            try:
                px = int(to)
            except Exception:
                raise ValueError(f"scroll 'to' must be 'bottom'|'top'|<int>, got: {to!r}")
            await page.evaluate(f"() => window.scrollTo(0, {px})")
    elif action == "wait_for_selector":
        await page.wait_for_selector(str(arg), timeout=bounded(None, 15_000))
    elif action == "wait_for_load_state":
        await page.wait_for_load_state(str(arg or "networkidle"),
                                          timeout=bounded(None, 15_000))
    elif action == "wait_ms":
        await page.wait_for_timeout(int(arg or 0))
    elif action == "screenshot":
        # Intermediate screenshot — captured to logs but not returned
        # (keeping the MCP response small). The act() caller still gets
        # the FINAL screenshot fed into Sonnet vision.
        name = (arg or {}).get("name", "step") if isinstance(arg, dict) else "step"
        png = await page.screenshot(type="png", full_page=False)
        log.info(json.dumps({"evt": "step_screenshot", "name": name,
                              "bytes": len(png)}))
    else:
        raise ValueError(f"unsupported action: {action!r}")


# ============================================================================
# API discovery + replay. The pattern these implement, end to end:
#   1. inspect_network(url[, steps])  → see which XHR/fetch the page fires and
#      with what params (the discovery step).
#   2. call_api(endpoint, method, body) → replay that endpoint directly, with
#      arbitrary params, from the page's own origin (the replay step). This
#      reaches periods the UI never exposes (e.g. a year missing from a
#      dropdown) and never touches the brittle widget.
#   3. Persist the recipe as a playbook `api` entry so step 1 is skipped next
#      time.
# ============================================================================

def _encode_body(body: Any, content_type: str | None) -> tuple[Any, str | None]:
    """Turn a dict/str body into (wire_string, content_type). A dict defaults to
    form-encoding (what most gov AJAX endpoints want); pass content_type
    'application/json' to send it as JSON instead."""
    if body is None:
        return None, content_type
    if isinstance(body, str):
        return body, content_type
    if isinstance(body, dict):
        if content_type and "json" in content_type:
            return json.dumps(body), content_type
        from urllib.parse import urlencode
        return urlencode(body), (content_type
                                 or "application/x-www-form-urlencoded")
    return str(body), content_type


async def _page_fetch(page: Page, url: str, *, method: str = "GET",
                      body: Any = None, headers: dict[str, str] | None = None,
                      content_type: str | None = None,
                      body_cap: int = 40_000) -> dict[str, Any]:
    """Run fetch() inside the page so the request inherits the page's origin,
    cookies and session. Returns {status, content_type, ok, json|text}."""
    wire_body, ct = _encode_body(body, content_type)
    hdrs = {"X-Requested-With": "XMLHttpRequest"}
    if ct:
        hdrs["Content-Type"] = ct
    if headers:
        hdrs.update(headers)
    method = method.upper()
    res = await page.evaluate(
        """async (a) => {
            const opt = {method: a.method, headers: a.headers,
                         credentials: 'include'};
            if (a.body != null && a.method !== 'GET' && a.method !== 'HEAD')
                opt.body = a.body;
            const r = await fetch(a.url, opt);
            const text = await r.text();
            return {status: r.status, ok: r.ok,
                    content_type: r.headers.get('content-type') || '', text};
        }""",
        {"url": url, "method": method, "headers": hdrs, "body": wire_body},
    )
    text = res.get("text") or ""
    out: dict[str, Any] = {
        "status": res.get("status"),
        "ok": res.get("ok"),
        "content_type": res.get("content_type", ""),
    }
    parsed = None
    if text:
        try:
            parsed = json.loads(text)
        except Exception:  # noqa: BLE001
            parsed = None
    if parsed is not None:
        out["json"] = parsed
    else:
        out["text"] = text[:body_cap]
        out["text_truncated"] = len(text) > body_cap
    return out


async def call_api(
    url: str,
    *,
    method: str = "GET",
    body: Any = None,
    headers: dict[str, str] | None = None,
    page_url: str | None = None,
    content_type: str | None = None,
    timeout_ms: int = 30_000,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Replay an API/AJAX endpoint directly from a real browser origin.

    Loads a page on the endpoint's origin first (so cookies, CSRF state,
    Origin/Referer all match), then issues the request via in-page fetch().
    Bypasses brittle UI widgets entirely and reaches data the front-end never
    surfaces (e.g. a fiscal year missing from a dropdown).

    Args:
        url: The endpoint URL (absolute).
        method: HTTP method (GET/POST/…). Default GET.
        body: Request body — a dict (form-encoded by default; JSON if
            content_type is application/json) or a pre-encoded string.
        headers: Extra request headers (merged over the XHR defaults).
        page_url: Origin page to load before fetching. Defaults to the
            endpoint's scheme://host/. Set this to the actual dashboard URL
            when the endpoint checks Referer.
        content_type: Override the request Content-Type.
        timeout_ms: Navigation timeout for loading page_url.

    Returns:
        {url, page_url, status, content_type, ok, json|text, source: "browser_api"}
    """
    t0 = time.perf_counter()
    if not url:
        return {"error": "url is required"}
    parsed_u = urlparse(url)
    origin = (page_url or
              f"{parsed_u.scheme or 'https'}://{parsed_u.netloc}/")
    client_id = _current_client.get() or "anon"
    ctx = await _get_context(client_id, proxied=use_proxy)
    page = await ctx.new_page()
    try:
        try:
            await page.goto(origin, wait_until="domcontentloaded",
                            timeout=timeout_ms)
        except Exception as e:  # noqa: BLE001
            _emit("call_api", t0, extra={"error": f"goto:{str(e)[:60]}"})
            return {"error": f"could not load origin page {origin}: {e}",
                    "url": url, "page_url": origin}
        try:
            res = await _page_fetch(page, url, method=method, body=body,
                                    headers=headers, content_type=content_type)
        except Exception as e:  # noqa: BLE001
            _emit("call_api", t0, extra={"error": f"fetch:{str(e)[:60]}"})
            return {"error": f"in-page fetch failed: {e}", "url": url,
                    "page_url": origin}
        out = {"url": url, "page_url": origin, "domain": _domain(url),
               "source": "browser_api", **res,
               "fetched_at": datetime.now(timezone.utc).isoformat()}
        _emit("call_api", t0, extra={"status": res.get("status"),
                                     "has_json": "json" in res})
        return out
    finally:
        try:
            await page.close()
        except Exception:  # noqa: BLE001
            pass


async def inspect_network(
    url: str,
    *,
    steps: list[dict[str, Any]] | None = None,
    timeout_ms: int = 60_000,
    settle_ms: int = 2500,
    url_filter: str | None = None,
    max_entries: int = 80,
    body_cap: int = 4000,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Open a page (optionally running `act`-style steps) and report the
    XHR/fetch calls it fires — endpoint, method, request params, response sample.

    This is the discovery half of the API-replay pattern. Run it once on a
    JS-driven dashboard to learn which endpoint feeds the table/chart and what
    params it takes; then hit that endpoint with call_api (templating the params
    for the period you actually want). Pass `steps` to capture the request a
    dropdown/tab/button fires — e.g. select a year, then read the endpoint it hit.

    Args:
        url: Page to open.
        steps: Optional `act`-vocabulary steps to run while recording (e.g.
            change a dropdown so its AJAX call is captured).
        timeout_ms: Navigation/step timeout.
        settle_ms: Extra wait after load/steps so late XHRs are captured.
        url_filter: Optional substring; only requests whose URL contains it are
            returned (e.g. "Ajax" or "/api/").
        max_entries: Cap on recorded requests.
        body_cap: Per-response sample cap (chars).

    Returns:
        {url, final_url, request_count, requests: [{method, url,
         resource_type, status, content_type, request_params, response_sample,
         response_truncated}], step_results?}
    """
    t0 = time.perf_counter()
    if not url:
        return {"error": "url is required"}
    client_id = _current_client.get() or "anon"
    ctx = await _get_context(client_id, proxied=use_proxy)
    page = await ctx.new_page()
    rec = _NetworkRecorder(max_entries=max_entries, body_cap=body_cap)
    rec.attach(page)
    step_results: list[dict[str, Any]] = []
    try:
        try:
            await page.goto(url, wait_until="domcontentloaded",
                            timeout=timeout_ms)
        except Exception as e:  # noqa: BLE001
            _emit("inspect_network", t0, extra={"error": f"goto:{str(e)[:60]}"})
            return {"error": f"navigation failed: {e}", "url": url}
        try:
            await page.wait_for_load_state("networkidle", timeout=8_000)
        except Exception:  # noqa: BLE001
            pass
        for idx, step in enumerate(steps or []):
            if not isinstance(step, dict) or len(step) != 1:
                step_results.append({"step_index": idx, "ok": False,
                                     "error": "each step must be {action: arg}"})
                continue
            action, arg = next(iter(step.items()))
            s_t0 = time.perf_counter()
            try:
                await _run_step(page, action, arg, timeout_ms)
                step_results.append({"step_index": idx, "action": action,
                                     "ok": True,
                                     "duration_ms": round(
                                         (time.perf_counter() - s_t0) * 1000)})
            except Exception as e:  # noqa: BLE001
                step_results.append({"step_index": idx, "action": action,
                                     "ok": False, "error": str(e)[:200]})
        if settle_ms > 0:
            await page.wait_for_timeout(settle_ms)
        final_url = page.url
        await rec.finalize()

        def _shape(e: dict[str, Any]) -> dict[str, Any]:
            rb = e.get("request_body") or {}
            return {
                "method": e["method"],
                "url": e["url"],
                "resource_type": e["resource_type"],
                "status": e.get("status"),
                "content_type": e.get("content_type", ""),
                "request_params": rb.get("data") if isinstance(rb, dict) else None,
                "request_body_kind": rb.get("kind") if isinstance(rb, dict) else None,
                "response_sample": e.get("response_sample"),
                "response_truncated": e.get("response_truncated", False),
            }

        reqs = [_shape(e) for e in rec.discovered_api(limit=max_entries)]
        if url_filter:
            reqs = [r for r in reqs if url_filter.lower() in r["url"].lower()]
        out: dict[str, Any] = {
            "url": url,
            "final_url": final_url,
            "domain": _domain(final_url),
            "request_count": len(reqs),
            "requests": reqs,
            "fetched_at": datetime.now(timezone.utc).isoformat(),
        }
        if steps:
            out["step_results"] = step_results
        _emit("inspect_network", t0, extra={"reqs": len(reqs),
                                            "steps": len(steps or [])})
        return out
    finally:
        try:
            await page.close()
        except Exception:  # noqa: BLE001
            pass


# ============================================================================
# smart_fetch — playbook-AWARE fetch. Don't just attach the playbook; ACT on it:
# replay its `api` endpoint, pull its `open_data` mirror, else render. This is
# what the upstream web_fetch escalation calls, so a known site resolves via its
# recipe instead of a blind browser render.
# ============================================================================

_PLACEHOLDER_RE = re.compile(r"<[^>]*>")


def _derive_fy(text: str) -> str | None:
    """Pull an Indian fiscal year (YYYY-YYYY) out of free text — '2023-2024',
    '2023-24', 'FY24', 'FY2024' (FY label = the year it ENDS). Returns None when
    ambiguous (a bare 4-digit year) so callers fall back rather than guess."""
    t = text or ""
    m = re.search(r"(20\d\d)\s*[-/]\s*(20\d\d)", t)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(20\d\d)\s*[-/]\s*(\d\d)\b", t)
    if m:
        return f"{m.group(1)}-20{m.group(2)}"
    m = re.search(r"\bFY\s*(?:20)?(\d\d)\b", t, re.IGNORECASE)
    if m:
        end = 2000 + int(m.group(1))
        return f"{end - 1}-{end}"
    return None


def _template_api_params(params: dict[str, Any] | None,
                         focus: str) -> dict[str, Any] | None:
    """Fill <placeholder> values in a playbook `api` param template from `focus`.
    Today only fiscal-year-shaped params auto-fill; any other unresolved
    placeholder returns None so the caller falls back to a safer rung."""
    if not isinstance(params, dict):
        return None
    out: dict[str, Any] = {}
    fy: str | None = None
    for k, v in params.items():
        if isinstance(v, str) and _PLACEHOLDER_RE.search(v):
            kl = k.lower()
            if "year" in kl or "fy" in kl or "financial" in kl:
                fy = fy or _derive_fy(focus)
                if not fy:
                    return None
                out[k] = fy
            else:
                return None
        else:
            out[k] = v
    return out


async def smart_fetch(url: str, *, focus: str = "",
                      use_proxy: bool = False) -> dict[str, Any]:
    """Playbook-aware fetch: consult the URL's playbook and ACT on it — replay
    its `api` endpoint (templating params from `focus`), or pull its `open_data`
    mirror — falling back to a full browser render (`extract`) when there's no
    auto-executable recipe. Returns the `extract` structured shape plus
    `rung_used` (api|open_data|render) and `playbook_id`."""
    t0 = time.perf_counter()
    if not url:
        return {"error": "url is required"}
    from . import playbooks as _pb
    try:
        pb = await _pb.match_for_url(url)
    except Exception:  # noqa: BLE001
        pb = None

    if pb:
        # (a) api recipe → call_api → structured extract over the JSON.
        for api in (pb.get("api") or []):
            endpoint = api.get("endpoint")
            if not endpoint or api.get("params") is None:
                continue
            templated = _template_api_params(api.get("params"), focus)
            if templated is None:
                continue
            try:
                r = await call_api(endpoint, method=api.get("method", "GET"),
                                   body=templated or None, page_url=url,
                                   use_proxy=use_proxy)
            except Exception:  # noqa: BLE001
                continue
            data = r.get("json") if isinstance(r, dict) else None
            if data is not None:
                out = await _sonnet_extract(
                    {"url": url, "domain": _domain(url),
                     "text": json.dumps(data, default=str)[:16000],
                     "fetched_at": datetime.now(timezone.utc).isoformat()},
                    focus=focus)
                out.update(rung_used="api", playbook_id=pb.get("id"),
                           api_endpoint=endpoint, api_params=templated)
                _emit("smart_fetch", t0, extra={"rung": "api"})
                return out
        # (b) open_data: a concrete download_file URL (open mirror / auth pivot).
        for od in (pb.get("open_data") or []):
            u = od.get("url")
            if u and (od.get("tool") or "download_file") == "download_file":
                try:
                    r = await download_file(u, query=(focus or None),
                                            use_proxy=use_proxy)
                except Exception:  # noqa: BLE001
                    continue
                if isinstance(r, dict) and "error" not in r:
                    r.update(rung_used="open_data", playbook_id=pb.get("id"))
                    _emit("smart_fetch", t0, extra={"rung": "open_data"})
                    return r

    # (c) default → full browser render.
    out = await extract(url, focus=focus, use_proxy=use_proxy)
    if isinstance(out, dict):
        out["rung_used"] = "render"
        if pb:
            out.setdefault("playbook_id", pb.get("id"))
    _emit("smart_fetch", t0, extra={"rung": "render"})
    return out


async def _sonnet_extract(visited: dict[str, Any], *, focus: str = "") -> dict[str, Any]:
    """Shared LLM extraction over a `visited`-shaped payload. Used by extract()
    and act() so both produce identical output schemas."""
    base = {
        "url": visited["url"],
        "domain": visited.get("domain", ""),
        "fetched_at": visited["fetched_at"],
        "kind": "browser",
        # Provenance: "browser" (Chromium), "tavily", or "anthropic_web_fetch".
        "source": visited.get("source") or "browser",
    }
    if visited.get("blocked"):
        base["blocked"] = visited["blocked"]
    text = visited.get("text") or ""
    shot_b64 = visited.get("screenshot_b64")

    client = await _anthropic()
    if client is None:
        return {
            **base, "title": visited.get("title", ""),
            "dateline": "", "summary": text[:400],
            "key_facts": [], "numeric_values": [], "dates": [],
            "tables_summary": [], "raw_content": text[:8000],
            "note": "ANTHROPIC_API_KEY not set — returning raw content only.",
        }

    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    user_content: list[dict[str, Any]] = []
    if shot_b64:
        user_content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png",
                         "data": shot_b64},
        })
    user_content.append({
        "type": "text",
        "text": (
            f"FOCUS: {focus or '(general — extract the main facts and tables)'}\n\n"
            f"PAGE TITLE: {visited.get('title', '')}\n"
            f"PAGE URL: {visited.get('url', '')}\n\n"
            f"PAGE TEXT (innerText, truncated):\n{text[:16000]}"
        ),
    })

    try:
        resp = await client.messages.create(
            model=_anthropic_model(),
            max_tokens=3000,
            system=[
                {"type": "text",
                 "text": STRUCTURED_EXTRACT_SYSTEM_STATIC,
                 "cache_control": {"type": "ephemeral"}},
                {"type": "text",
                 "text": dynamic_date_block(today_iso)},
            ],
            messages=[{"role": "user", "content": user_content}],
        )
        body = "".join(b.text for b in resp.content if b.type == "text").strip()
    except Exception as e:  # noqa: BLE001
        log.warning("act/extract LLM call failed: %s", e)
        return {**base, "title": visited.get("title", ""),
                "error_extract": str(e)[:200], "raw_content": text[:8000]}

    body = re.sub(r"^```(?:json)?\s*", "", body)
    body = re.sub(r"\s*```$", "", body)
    parsed = _parse_relaxed_json(body)

    out = {
        **base, "focus": focus,
        "title": parsed.get("title", visited.get("title", "")) or visited.get("title", ""),
        "dateline": parsed.get("dateline", ""),
        "summary": parsed.get("summary", ""),
        "key_facts": parsed.get("key_facts", []) or [],
        "numeric_values": parsed.get("numeric_values", []) or [],
        "dates": parsed.get("dates", []) or [],
        "tables_summary": parsed.get("tables_summary", []) or [],
    }
    if not parsed:
        out["raw_content"] = text[:8000]
        out["note"] = "Structured extraction failed; returning raw content."
    if shot_b64:
        out["screenshot_bytes"] = visited.get("screenshot_bytes")
    return out


async def extract(
    url: str,
    *,
    focus: str = "",
    wait_for_selector: str | None = None,
    full_page_screenshot: bool = True,
    include_screenshot_in_response: bool = False,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Visit URL with Chromium → Sonnet structured extraction.

    Sends both the rendered text AND a full-page screenshot to Sonnet, so
    numbers drawn via canvas / SVG (charts on PPAC, RBI, NSE dashboards)
    that don't appear in the DOM text still get picked up.

    Returns the same shape as authority-web-search's pdf_fetch_structured /
    web_fetch_structured: {title, dateline, summary, key_facts,
    numeric_values, dates, tables_summary}.
    """
    t0 = time.perf_counter()
    visited = await visit(
        url,
        wait_for_selector=wait_for_selector,
        screenshot=True,
        full_page_screenshot=full_page_screenshot,
        text_cap=20_000,
        # We NEED the bytes for Sonnet vision — extract() always asks for them.
        # _sonnet_extract reads `screenshot_b64`; the final response strips it
        # back out unless the caller explicitly asked for it via
        # include_screenshot_in_response.
        return_screenshot_b64=True,
        use_proxy=use_proxy,
    )
    if "error" in visited:
        return visited

    out = await _sonnet_extract(visited, focus=focus)
    if include_screenshot_in_response and visited.get("screenshot_b64"):
        out["screenshot_b64"] = visited["screenshot_b64"]
    _emit("extract", t0,
           extra={"shot_kb": round((visited.get("screenshot_bytes", 0)) / 1024)})
    return out


# ============================================================================
# download_file — fetch + parse spreadsheets and PDFs.
#
# Port of the battle-tested fetch+parse pipeline from authority-web-search.
# One tool, format auto-detected by content-type + magic bytes, parsed by
# openpyxl / xlrd / csv / pypdf, returns the same classified-error shape so
# the calling model can navigate identically across both MCPs.
#
# This closes the loop on URLs that the Chromium-based visit/act/extract
# tools surface as `file_links` but can't read themselves — gov-site Excel
# bulletins (gst.gov.in, cga.nic.in, mospi.gov.in), PDF circulars (RBI,
# SEBI), CSV data dumps (data.gov.in).
# ============================================================================

_EXCEL_DOWNLOAD_CAP_BYTES = 16 * 1024 * 1024   # 16 MB
_PDF_DOWNLOAD_CAP_BYTES = 24 * 1024 * 1024     # 24 MB
_EXCEL_TEXT_CAP = 80_000                       # chars
_PDF_TEXT_CAP = 80_000                         # chars


def _is_excel_url(url: str) -> bool:
    if not url:
        return False
    try:
        u = urlparse(url.lower())
    except Exception:
        return False
    for ext in (".xlsx", ".xlsm", ".xls", ".csv", ".tsv"):
        if u.path.endswith(ext) or ext in (u.query or ""):
            return True
    return False


def _is_pdf_url(url: str) -> bool:
    if not url:
        return False
    try:
        u = urlparse(url.lower())
    except Exception:
        return False
    return u.path.endswith(".pdf") or ".pdf" in (u.query or "")


def _parse_excel_sync(raw_bytes: bytes, sheet: str | None,
                       max_rows_per_sheet: int) -> dict[str, Any]:
    import io as _io
    sheets: list[dict[str, Any]] = []
    fmt = "xlsx"
    try:
        from openpyxl import load_workbook
    except Exception as e:  # noqa: BLE001
        return {"error": f"openpyxl unavailable: {e}"}
    try:
        wb = load_workbook(filename=_io.BytesIO(raw_bytes), read_only=True,
                            data_only=True)
        sheet_names = wb.sheetnames
        targets = [sheet] if sheet and sheet in sheet_names else sheet_names
        chunks: list[str] = []
        total_chars = 0
        for sn in targets:
            ws = wb[sn]
            rows_out: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [("" if v is None else str(v)) for v in row]
                while vals and not vals[-1].strip():
                    vals.pop()
                if not vals:
                    continue
                rows_out.append(vals)
                if i >= max_rows_per_sheet:
                    break
            header = rows_out[0] if rows_out else []
            sample = rows_out[1:11]
            sheets.append({
                "name": sn,
                "rows": len(rows_out),
                "cols": max((len(r) for r in rows_out), default=0),
                "header": header,
                "sample": sample,
            })
            chunk = f"\n\n=== Sheet: {sn} ===\n" + "\n".join(
                "\t".join(r) for r in rows_out
            )
            if total_chars + len(chunk) > _EXCEL_TEXT_CAP:
                chunk = chunk[: _EXCEL_TEXT_CAP - total_chars]
                chunks.append(chunk)
                break
            chunks.append(chunk)
            total_chars += len(chunk)
        return {"sheets": sheets, "content": "".join(chunks).strip(),
                "format": fmt, "sheet_count": len(sheet_names)}
    except Exception as e:  # noqa: BLE001
        # Fall back to xlrd for legacy .xls before giving up.
        try:
            import xlrd  # type: ignore
            book = xlrd.open_workbook(file_contents=raw_bytes)
            chunks: list[str] = []
            total_chars = 0
            for s_idx in range(book.nsheets):
                ws = book.sheet_by_index(s_idx)
                rows_out: list[list[str]] = []
                for i in range(ws.nrows):
                    vals = [str(ws.cell_value(i, j)) for j in range(ws.ncols)]
                    while vals and not vals[-1].strip():
                        vals.pop()
                    if not vals:
                        continue
                    rows_out.append(vals)
                    if i >= max_rows_per_sheet:
                        break
                header = rows_out[0] if rows_out else []
                sample = rows_out[1:11]
                sheets.append({
                    "name": ws.name, "rows": len(rows_out),
                    "cols": max((len(r) for r in rows_out), default=0),
                    "header": header, "sample": sample,
                })
                chunk = f"\n\n=== Sheet: {ws.name} ===\n" + "\n".join(
                    "\t".join(r) for r in rows_out
                )
                if total_chars + len(chunk) > _EXCEL_TEXT_CAP:
                    chunk = chunk[: _EXCEL_TEXT_CAP - total_chars]
                    chunks.append(chunk)
                    break
                chunks.append(chunk)
                total_chars += len(chunk)
            return {"sheets": sheets, "content": "".join(chunks).strip(),
                    "format": "xls", "sheet_count": book.nsheets}
        except Exception as e2:  # noqa: BLE001
            return {"error": f"could not open Excel: {e}; xls fallback: {e2}"}


def _parse_csv_sync(raw_bytes: bytes, max_rows: int) -> dict[str, Any]:
    import csv as _csv
    text = raw_bytes.decode("utf-8", errors="replace")
    try:
        dialect = _csv.Sniffer().sniff(text[:4096])
    except Exception:
        dialect = _csv.excel
    reader = _csv.reader(text.splitlines(), dialect=dialect)
    rows: list[list[str]] = []
    for i, r in enumerate(reader):
        rows.append([c.strip() for c in r])
        if i >= max_rows:
            break
    header = rows[0] if rows else []
    sample = rows[1:11]
    content = "\n".join("\t".join(r) for r in rows)[: _EXCEL_TEXT_CAP]
    return {"sheets": [{"name": "CSV", "rows": len(rows),
                          "cols": max((len(r) for r in rows), default=0),
                          "header": header, "sample": sample}],
            "content": content, "format": "csv", "sheet_count": 1}


def _parse_pdf_sync(raw_bytes: bytes, pages: list[int] | None,
                    max_pages: int) -> dict[str, Any]:
    import io as _io
    try:
        from pypdf import PdfReader
    except Exception as e:  # noqa: BLE001
        return {"error": f"pypdf unavailable: {e}"}
    try:
        reader = PdfReader(_io.BytesIO(raw_bytes))
        n_pages = len(reader.pages)
    except Exception as e:  # noqa: BLE001
        return {"error": f"could not open PDF: {e}"}
    wanted = ([i - 1 for i in pages if 1 <= i <= n_pages] if pages
              else list(range(min(n_pages, max_pages))))
    chunks: list[str] = []
    extracted: list[int] = []
    total_chars = 0
    truncated = False
    for i in wanted:
        try:
            txt = reader.pages[i].extract_text() or ""
        except Exception:  # noqa: BLE001
            txt = ""
        if not txt:
            continue
        chunk = f"\n\n--- Page {i + 1} ---\n{txt}"
        if total_chars + len(chunk) > _PDF_TEXT_CAP:
            chunk = chunk[: _PDF_TEXT_CAP - total_chars]
            chunks.append(chunk)
            extracted.append(i + 1)
            truncated = True
            break
        chunks.append(chunk)
        extracted.append(i + 1)
        total_chars += len(chunk)
    return {
        "content": "".join(chunks).strip(),
        "page_count": n_pages,
        "pages_extracted": extracted,
        "content_truncated": truncated or len(wanted) < n_pages,
    }


# ============================================================================
# Economical, query-targeted reading. Instead of dumping a whole PDF / workbook
# into context, scan it for the query terms and return only the matching pages
# (PDF) or rows (spreadsheet/CSV) with a snippet — pdfgrep, built in. This is how
# you answer "what's the April-2024 fiscal-deficit figure in this 200-page
# Monthly Accounts PDF" without paying for 200 pages. A page/row matches when it
# contains ALL whitespace-separated query tokens (case-insensitive).
# ============================================================================

def _query_tokens(query: str) -> list[str]:
    return [t for t in re.split(r"\s+", (query or "").lower().strip()) if t]


def _text_matches(text: str, tokens: list[str]) -> bool:
    if not tokens:
        return False
    low = text.lower()
    return all(t in low for t in tokens)


def _snippet(text: str, tokens: list[str], ctx: int = 240) -> str:
    """A grep -C style window around the earliest matching token."""
    low = text.lower()
    hits = [low.find(t) for t in tokens if low.find(t) >= 0]
    pos = min(hits) if hits else -1
    if pos < 0:
        return " ".join(text[:ctx].split())
    start, end = max(0, pos - ctx // 2), min(len(text), pos + ctx // 2)
    body = " ".join(text[start:end].split())
    return ("…" if start > 0 else "") + body + ("…" if end < len(text) else "")


def _grep_pdf_sync(raw_bytes: bytes, query: str, max_pages_scan: int = 400,
                   max_matches: int = 40) -> dict[str, Any]:
    import io as _io
    try:
        from pypdf import PdfReader
    except Exception as e:  # noqa: BLE001
        return {"error": f"pypdf unavailable: {e}"}
    try:
        reader = PdfReader(_io.BytesIO(raw_bytes))
        n = len(reader.pages)
    except Exception as e:  # noqa: BLE001
        return {"error": f"could not open PDF: {e}"}
    tokens = _query_tokens(query)
    matches: list[dict[str, Any]] = []
    scanned = 0
    for i in range(min(n, max_pages_scan)):
        scanned += 1
        try:
            txt = reader.pages[i].extract_text() or ""
        except Exception:  # noqa: BLE001
            txt = ""
        if txt and _text_matches(txt, tokens):
            matches.append({"page": i + 1, "snippet": _snippet(txt, tokens)})
            if len(matches) >= max_matches:
                break
    return {"matches": matches, "match_count": len(matches),
            "page_count": n, "pages_scanned": scanned}


def _grep_excel_sync(raw_bytes: bytes, query: str, sheet: str | None,
                     max_matches: int = 80) -> dict[str, Any]:
    import io as _io
    try:
        from openpyxl import load_workbook
    except Exception as e:  # noqa: BLE001
        return {"error": f"openpyxl unavailable: {e}"}
    try:
        wb = load_workbook(filename=_io.BytesIO(raw_bytes), read_only=True,
                           data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"error": f"could not open Excel: {e}"}
    tokens = _query_tokens(query)
    targets = [sheet] if sheet and sheet in wb.sheetnames else wb.sheetnames
    matches: list[dict[str, Any]] = []
    for sn in targets:
        ws = wb[sn]
        header: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [("" if v is None else str(v)) for v in row]
            if i == 0:
                header = vals
            if _text_matches("\t".join(vals), tokens):
                matches.append({"sheet": sn, "row_index": i + 1,
                                "header": header, "row": vals})
                if len(matches) >= max_matches:
                    return {"matches": matches, "match_count": len(matches)}
    return {"matches": matches, "match_count": len(matches)}


def _grep_csv_sync(raw_bytes: bytes, query: str,
                   max_matches: int = 80) -> dict[str, Any]:
    import csv as _csv
    tokens = _query_tokens(query)
    text = raw_bytes.decode("utf-8", errors="replace")
    try:
        dialect = _csv.Sniffer().sniff(text[:4096])
    except Exception:  # noqa: BLE001
        dialect = _csv.excel
    header: list[str] = []
    matches: list[dict[str, Any]] = []
    for i, row in enumerate(_csv.reader(text.splitlines(), dialect=dialect)):
        vals = [c.strip() for c in row]
        if i == 0:
            header = vals
        if _text_matches("\t".join(vals), tokens):
            matches.append({"sheet": "CSV", "row_index": i + 1,
                            "header": header, "row": vals})
            if len(matches) >= max_matches:
                break
    return {"matches": matches, "match_count": len(matches)}


async def download_file(
    url: str,
    *,
    sheet: str | None = None,
    pages: list[int] | None = None,
    max_rows_per_sheet: int = 200,
    max_pdf_pages: int = 30,
    query: str | None = None,
    max_matches: int = 40,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Download a URL and parse it as a spreadsheet (.xlsx/.xlsm/.xls/.csv/
    .tsv) or PDF. Format is auto-detected via content-type + magic bytes.

    ECONOMICAL READING — pass `query` to grep instead of dumping the whole
    file. The file is scanned for the query terms (ALL whitespace-separated
    tokens must appear, case-insensitive) and ONLY matching pages (PDF) or rows
    (xlsx/csv) come back, with a snippet. Use this for big Monthly-Accounts-style
    PDFs / workbooks where you want one figure, not 200 pages. Without `query`,
    behaviour is unchanged (full parse, capped).

    Returns one of:
      • Spreadsheet → {kind: "spreadsheet", url, domain, format,
                       sheets[], content, sheet_count, fetched_at}
      • PDF         → {kind: "pdf", url, domain, content, page_count,
                       pages_extracted, content_truncated, fetched_at}
      • PDF search  → {kind: "pdf_search", url, domain, query, matches:[{page,
                       snippet}], match_count, page_count, pages_scanned}
      • Sheet search→ {kind: "spreadsheet_search", url, domain, query,
                       matches:[{sheet, row_index, header, row}], match_count}
      • Error       → {error, error_kind, url, domain, ...}

    error_kind values: http_error, html_masquerade, truncated_body,
    invalid_xlsx, parse_error, wrong_content_type, too_large.
    """
    t0 = time.perf_counter()
    domain = _domain(url)
    # Stealth-ish headers so gov-CDNs that block "python-httpx/x.y" still
    # serve us. Same UA shape as the Playwright contexts.
    headers = {
        "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/127.0.0.0 Safari/537.36"),
        "Accept": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, "
                   "application/vnd.ms-excel, application/pdf, text/csv, */*;q=0.8"),
        "Accept-Language": "en-US,en;q=0.9",
    }
    proxy_url = _httpx_proxy_url() if use_proxy else None
    try:
        async with httpx.AsyncClient(timeout=90.0, follow_redirects=True,
                                       headers=headers, proxy=proxy_url) as client:
            r = await client.get(url)
    except httpx.HTTPError as e:
        out = {"error": f"download failed: {e}", "error_kind": "http_error",
                "url": url, "domain": domain}
        _emit("download_file", t0, extra={"err": "transport"})
        return out

    ct = (r.headers.get("content-type") or "").lower()
    body_head = r.content[:512]
    looks_like_html = (
        body_head[:5].lower() in (b"<html", b"<!doc")
        or b"<html" in body_head[:200].lower()
    )

    if r.status_code >= 400:
        _emit("download_file", t0, extra={"err": "http", "status": r.status_code})
        return {
            "error": (f"HTTP {r.status_code} from server — the URL is broken "
                       f"or you don't have access. NOT a parsing issue; pick a "
                       f"different file."),
            "error_kind": "http_error",
            "http_status": r.status_code,
            "url": url, "domain": domain,
        }
    if looks_like_html and not _is_pdf_url(url):
        _emit("download_file", t0, extra={"err": "html"})
        return {
            "error": ("Server returned an HTML page (likely a 404 redirect "
                       "or login wall), not a file. The URL probably "
                       "redirected somewhere else."),
            "error_kind": "html_masquerade",
            "url": url, "domain": domain,
            "body_head": body_head[:200].decode("utf-8", errors="replace"),
        }
    if len(r.content) < 256:
        _emit("download_file", t0, extra={"err": "truncated"})
        return {
            "error": (f"Server returned only {len(r.content)} bytes — too "
                       f"small to be a real file. The URL is almost certainly "
                       f"broken."),
            "error_kind": "truncated_body",
            "url": url, "domain": domain,
            "body_head": body_head.decode("utf-8", errors="replace"),
        }

    # Magic-byte detection: xlsx = PK (zip); xls = OLE; pdf = %PDF-.
    is_xlsx = r.content[:2] == b"PK"
    is_xls = r.content[:8].startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")
    is_pdf = (r.content[:5] == b"%PDF-"
              or r.content[:8].lstrip().startswith(b"%PDF-"))
    is_csv = (
        ("csv" in ct) or url.lower().endswith(".csv")
        or (not is_xlsx and not is_xls and not is_pdf
            and url.lower().endswith(".tsv"))
    )

    if is_pdf or "pdf" in ct or _is_pdf_url(url):
        if len(r.content) > _PDF_DOWNLOAD_CAP_BYTES:
            _emit("download_file", t0, extra={"err": "too_large",
                                                "bytes": len(r.content)})
            return {
                "error": (f"PDF too large ({len(r.content) // (1024*1024)} MB, "
                           f"cap {_PDF_DOWNLOAD_CAP_BYTES // (1024*1024)} MB)"),
                "error_kind": "too_large",
                "url": url, "domain": domain,
            }
        if query:
            g = await asyncio.to_thread(_grep_pdf_sync, r.content, query,
                                         max_matches=max_matches)
            if "error" in g:
                _emit("download_file", t0, extra={"err": "pdf_grep"})
                return {**g, "error_kind": "parse_error",
                        "url": url, "domain": domain}
            _emit("download_file", t0, extra={"fmt": "pdf_search",
                                               "matches": g["match_count"]})
            return {"kind": "pdf_search", "url": url, "domain": domain,
                    "query": query, **g,
                    "fetched_at": datetime.now(timezone.utc).isoformat()}
        parsed = await asyncio.to_thread(_parse_pdf_sync, r.content,
                                           pages, max_pdf_pages)
        if "error" in parsed:
            _emit("download_file", t0, extra={"err": "pdf_parse"})
            return {**parsed, "error_kind": "parse_error",
                    "url": url, "domain": domain}
        out = {
            "kind": "pdf",
            "url": url, "domain": domain,
            "content": parsed["content"],
            "page_count": parsed["page_count"],
            "pages_extracted": parsed["pages_extracted"],
            "content_truncated": parsed["content_truncated"],
            "fetched_at": datetime.now(timezone.utc).isoformat(),
        }
        _emit("download_file", t0, extra={"fmt": "pdf",
                                            "pages": parsed["page_count"]})
        return out

    if len(r.content) > _EXCEL_DOWNLOAD_CAP_BYTES:
        _emit("download_file", t0, extra={"err": "too_large",
                                            "bytes": len(r.content)})
        return {
            "error": (f"spreadsheet too large "
                       f"({len(r.content) // (1024*1024)} MB, cap "
                       f"{_EXCEL_DOWNLOAD_CAP_BYTES // (1024*1024)} MB)"),
            "error_kind": "too_large",
            "url": url, "domain": domain,
        }

    if query and (is_csv or is_xlsx or is_xls or _is_excel_url(url)
                  or "excel" in ct or "spreadsheet" in ct):
        if is_csv:
            g = await asyncio.to_thread(_grep_csv_sync, r.content, query,
                                         max_matches)
        else:
            g = await asyncio.to_thread(_grep_excel_sync, r.content, query,
                                         sheet, max_matches)
        if "error" in g:
            _emit("download_file", t0, extra={"err": "sheet_grep"})
            return {**g, "error_kind": "parse_error",
                    "url": url, "domain": domain}
        _emit("download_file", t0, extra={"fmt": "spreadsheet_search",
                                           "matches": g["match_count"]})
        return {"kind": "spreadsheet_search", "url": url, "domain": domain,
                "query": query, **g,
                "fetched_at": datetime.now(timezone.utc).isoformat()}

    if is_csv:
        parsed = await asyncio.to_thread(_parse_csv_sync, r.content,
                                           max_rows_per_sheet)
    elif (is_xlsx or is_xls or _is_excel_url(url)
            or "excel" in ct or "spreadsheet" in ct):
        parsed = await asyncio.to_thread(_parse_excel_sync, r.content,
                                           sheet, max_rows_per_sheet)
    else:
        _emit("download_file", t0, extra={"err": "wrong_ct", "ct": ct})
        return {"error": f"not a supported file format (content-type: "
                          f"{ct or 'unknown'}). download_file handles "
                          f".xlsx/.xls/.csv/.tsv/.pdf.",
                "error_kind": "wrong_content_type",
                "url": url, "domain": domain}

    if "error" in parsed:
        msg = parsed["error"]
        if "[Content_Types].xml" in msg or "Unknown ZIP file" in msg:
            _emit("download_file", t0, extra={"err": "invalid_xlsx"})
            return {
                "error": ("File downloaded but is not a valid xlsx — the URL "
                           "probably points to a corrupt or stub file. (Parse "
                           f"error: {msg[:120]})"),
                "error_kind": "invalid_xlsx",
                "url": url, "domain": domain,
            }
        _emit("download_file", t0, extra={"err": "parse"})
        return {**parsed, "error_kind": "parse_error",
                "url": url, "domain": domain}

    out = {
        "kind": "spreadsheet",
        "url": url, "domain": domain,
        "format": parsed["format"],
        "sheet_count": parsed["sheet_count"],
        "sheets": parsed["sheets"],
        "content": parsed["content"],
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    }
    _emit("download_file", t0, extra={"fmt": parsed["format"],
                                        "sheets": parsed["sheet_count"]})
    return out


# ============================================================================
# sitemap_probe — the cheapest discovery step, run BEFORE driving a page.
# robots.txt + sitemap.xml often hand you the stable data URLs (JSON/CSV/XLSX
# endpoints, the full URL inventory) directly, so you can download_file /
# call_api them and skip the browser entirely.
# ============================================================================

_SITEMAP_FETCH_CAP_BYTES = 5 * 1024 * 1024   # 5 MB per sitemap document
# URL shapes that usually point at fetchable data rather than an HTML page.
_DATA_LIKE_EXT = (".json", ".csv", ".tsv", ".xlsx", ".xls", ".xml", ".pdf")


def _parse_robots(text: str) -> dict[str, Any]:
    """Pull Sitemap: and Disallow: directives out of a robots.txt body."""
    sitemaps: list[str] = []
    disallow: list[str] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or ":" not in line:
            continue
        k, v = line.split(":", 1)
        k, v = k.strip().lower(), v.strip()
        if k == "sitemap" and v:
            sitemaps.append(v)
        elif k == "disallow" and v:
            disallow.append(v)
    return {"sitemaps": sitemaps, "disallow": disallow[:50]}


def _parse_sitemap_xml(text: str) -> list[str]:
    """Return every <loc> URL — works for both a <urlset> and a
    <sitemapindex> (the caller decides which by whether the locs are .xml)."""
    return [m.strip() for m in
            re.findall(r"<loc>\s*([^<\s]+)\s*</loc>", text, flags=re.IGNORECASE)]


def _decode_sitemap_bytes(raw: bytes, url: str) -> str:
    """Decode a sitemap body to text, transparently gunzipping a .xml.gz (by
    extension or gzip magic bytes) — httpx only auto-decompresses transport
    Content-Encoding, not a gzipped sitemap *file*."""
    if raw[:2] == b"\x1f\x8b" or url.lower().endswith(".gz"):
        try:
            import gzip as _gz
            raw = _gz.decompress(raw)
        except Exception:  # noqa: BLE001
            return ""
    return raw.decode("utf-8", errors="replace")


def _is_html_doc(text: str) -> bool:
    """True if the body is an HTML page (e.g. an SPA app-shell served at
    /sitemap.xml) rather than real sitemap XML — so the caller can say so
    instead of silently returning zero URLs."""
    head = text.lstrip()[:200].lower()
    return (head.startswith("<!doctype html") or head.startswith("<html")
            or "<html" in head)


def _is_data_like(url: str) -> bool:
    low = url.lower()
    return low.endswith(_DATA_LIKE_EXT) or "/api/" in low or low.endswith("/api")


async def sitemap_probe(
    url: str,
    *,
    max_urls: int = 200,
    max_sitemaps: int = 6,
    use_proxy: bool = False,
) -> dict[str, Any]:
    """Read robots.txt + sitemap(s) for a site and surface its URL inventory,
    highlighting data-like endpoints (.json/.csv/.xlsx/.xml/.pdf, /api/).

    Returns:
        {origin, robots_found, robots: {sitemaps[], disallow[]},
         sitemaps_fetched[], url_count, data_like_urls[], urls[], notes}
    """
    t0 = time.perf_counter()
    if not url:
        return {"error": "url is required"}
    pu = urlparse(url)
    origin = f"{pu.scheme or 'https'}://{pu.netloc}"
    headers = {
        "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/127.0.0.0 Safari/537.36"),
        "Accept": "text/plain, application/xml, text/xml, */*;q=0.8",
    }
    proxy_url = _httpx_proxy_url() if use_proxy else None
    notes: list[str] = []
    robots: dict[str, Any] = {"sitemaps": [], "disallow": []}
    robots_found = False
    sitemap_queue: list[str] = []
    seen_sitemaps: set[str] = set()
    page_urls: list[str] = []
    fetched: list[str] = []

    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True,
                                  headers=headers, proxy=proxy_url) as client:
        # 1. robots.txt → Sitemap: directives (+ disallow as context).
        try:
            rr = await client.get(f"{origin}/robots.txt")
            if rr.status_code < 400 and rr.text and "<html" not in \
                    rr.text[:200].lower():
                robots_found = True
                robots = _parse_robots(rr.text)
                sitemap_queue.extend(robots["sitemaps"])
        except httpx.HTTPError as e:
            notes.append(f"robots.txt fetch failed: {str(e)[:80]}")
        # 2. Fall back to the conventional location if robots named none.
        if not sitemap_queue:
            sitemap_queue.append(f"{origin}/sitemap.xml")

        # 3. Walk sitemaps (resolving one level of <sitemapindex>), bounded.
        while sitemap_queue and len(fetched) < max_sitemaps:
            sm = sitemap_queue.pop(0)
            if sm in seen_sitemaps:
                continue
            seen_sitemaps.add(sm)
            try:
                sr = await client.get(sm)
            except httpx.HTTPError as e:
                notes.append(f"sitemap fetch failed ({sm}): {str(e)[:60]}")
                continue
            if sr.status_code >= 400:
                notes.append(f"sitemap {sm} → HTTP {sr.status_code}")
                continue
            body = _decode_sitemap_bytes(sr.content, sm)
            if _is_html_doc(body):
                notes.append(f"{sm} returned HTML, not XML — no usable sitemap "
                             "here; use visit/file_links or call_api instead")
                continue
            body = body[:_SITEMAP_FETCH_CAP_BYTES]
            fetched.append(sm)
            locs = _parse_sitemap_xml(body)
            # Child sitemaps (a sitemapindex) end in .xml/.xml.gz; queue them.
            for loc in locs:
                low = loc.lower()
                if low.endswith(".xml") or low.endswith(".xml.gz"):
                    if (loc not in seen_sitemaps
                            and len(fetched) + len(sitemap_queue) < max_sitemaps):
                        sitemap_queue.append(loc)
                else:
                    page_urls.append(loc)
            if len(page_urls) >= max_urls * 4:
                notes.append("URL inventory truncated (large sitemap)")
                break

    # Dedup, cap, classify.
    deduped: list[str] = list(dict.fromkeys(page_urls))
    data_like = [u for u in deduped if _is_data_like(u)][: max_urls]
    out = {
        "origin": origin,
        "robots_found": robots_found,
        "robots": robots,
        "sitemaps_fetched": fetched,
        "url_count": len(deduped),
        "data_like_urls": data_like,
        "urls": deduped[:max_urls],
        "notes": notes,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    }
    if data_like:
        out["hint"] = ("data-like URLs found — download_file (.xlsx/.csv/.pdf) "
                       "or call_api (/api, .json) these directly; skip the "
                       "browser.")
    _emit("sitemap_probe", t0, extra={"urls": len(deduped),
                                       "data_like": len(data_like),
                                       "sitemaps": len(fetched)})
    return out
